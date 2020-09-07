from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet('sheet_test2')

ws['A1'] = 'alghost'
ws['B1'] = 'test'

wb.save('result.xlsx')