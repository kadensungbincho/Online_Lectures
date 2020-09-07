from openpyxl import load_workbook

wb = load_workbook('data.xlsx', read_only=True)
data = wb.active

for row in data.iter_rows():
    print(row)